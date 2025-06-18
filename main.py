from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import Dict
import fitz  # PyMuPDF
import os
import uuid
import json
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill

app = FastAPI()

# CORS para frontend en GitHub Pages
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://jhowmsm.github.io"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.post("/procesar/")
async def procesar_pdf(
    file: UploadFile = File(...),
    referencias: str = Form(...),
    exclusiones: str = Form("{}")
):
    try:
        referencias = json.loads(referencias)
        exclusiones = json.loads(exclusiones)
    except Exception:
        return {"error": "No se pudo leer referencias o exclusiones"}

    columna_x = {}
    resultados = {ref: [] for ref in referencias}
    filename = file.filename
    temp_pdf = f"/tmp/{filename}"

    with open(temp_pdf, "wb") as f:
        f.write(await file.read())

    # Buscar NIE/NIF globalmente en todo el PDF
    nie_patron = re.compile(r"[XY]\d{7}W")
    nie_encontrados = set()
    with fitz.open(temp_pdf) as doc:
        for pagina in doc:
            bloques = pagina.get_text("dict")["blocks"]
            for bloque in bloques:
                for linea in bloque.get("lines", []):
                    for span in linea.get("spans", []):
                        texto = span["text"].strip()
                        codigos = nie_patron.findall(texto)
                        for codigo in codigos:
                            nie_encontrados.add(codigo)

    # Reiniciar para extraer columnas por referencia
    with fitz.open(temp_pdf) as doc:
        coordenadas_columnas = {}
        coordenadas_textos = {ref: [] for ref in referencias}

        for pagina in doc:
            bloques = pagina.get_text("dict")["blocks"]
            for bloque in bloques:
                for linea in bloque.get("lines", []):
                    for span in linea.get("spans", []):
                        texto = span["text"].strip()
                        x = span["bbox"][0]
                        ancho = span["bbox"][2] - span["bbox"][0]

                        for ref in referencias:
                            if ref in texto and ref not in columna_x:
                                columna_x[ref] = x
                            if ref in columna_x and abs(x - columna_x[ref]) <= 5:
                                if texto not in exclusiones.get(ref, []):
                                    resultados[ref].append(texto)
                                    coordenadas_textos[ref].append(ancho)

    max_len = max(len(vals) for vals in resultados.values())
    for ref in referencias:
        while len(resultados[ref]) < max_len:
            resultados[ref].append("")
            coordenadas_textos[ref].append(0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    ws.append(referencias)

    amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for i in range(max_len):
        fila = []
        for ref in referencias:
            fila.append(resultados[ref][i])
        ws.append(fila)

        # Aplicar color solo si columna es "Nombre/Razón Social" y ancho sospechoso
        for col_idx, ref in enumerate(referencias):
            if ref == "Nombre/Razón Social" and coordenadas_textos[ref][i] >= 190:
                ws.cell(row=ws.max_row, column=col_idx + 1).fill = amarillo

    if nie_encontrados:
        ws_alerta = wb.create_sheet("NIE_Warnings")
        ws_alerta.append(["NIE/NIF Detectados"])
        for codigo in sorted(nie_encontrados):
            ws_alerta.append([codigo])

    output_path = f"/tmp/resultado_{uuid.uuid4().hex}.xlsx"
    wb.save(output_path)

    return FileResponse(
        output_path,
        filename="resultado.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
